#!/usr/bin/env python3
"""
Bot 606 DGII — Telegram con flujo guiado por botones.

Acceso:
  El bot es privado. Para empezar hay que enviar la contraseña (BOT_PASSWORD,
  por defecto "/Juan2202"). Hasta entonces no responde a nada más.

Flujo de captura:
  1. /nueva (o enviar foto directamente)
  2. ¿A qué mes (Excel) va? → [Abr] [May] [Jun] … [Auto por fecha]
  3. ¿Dónde fue la compra?  → [Punta Cana] [Santo Domingo]
  4. ¿Para qué es?          → [Casa] [Obra]
  5. Envía la foto de la factura
  6. Revisión de datos      → [✅ Aceptar] [✏️ Corregir] [❌ Cancelar]
     └ Corregir:            → seleccionar campo → escribir nuevo valor → volver a revisión

Otros comandos:
  /resumen   — totales del mes
  /lista     — todas las facturas
  /pendientes — facturas con advertencias
  /exportar  — descargar Excel 606
  /borrar    — eliminar última factura
  /mes YYYY-MM — cambiar mes activo
"""

import asyncio
import base64
import io
import json
import logging
import os
import re
import sqlite3
from datetime import datetime
from urllib.parse import parse_qs, urlparse

import anthropic
from dotenv import load_dotenv

import drive_sync
from telegram import (
    InlineKeyboardButton,
    InlineKeyboardMarkup,
    KeyboardButton,
    ReplyKeyboardMarkup,
    Update,
)
from telegram.ext import (
    Application,
    ApplicationHandlerStop,
    CallbackQueryHandler,
    CommandHandler,
    ContextTypes,
    ConversationHandler,
    MessageHandler,
    TypeHandler,
    filters,
)

load_dotenv()

logging.basicConfig(
    format="%(asctime)s [%(levelname)s] %(message)s",
    level=logging.INFO,
)
log = logging.getLogger(__name__)

# ──────────────────────────────────────────────────────────────
# CONFIG
# ──────────────────────────────────────────────────────────────

BOT_TOKEN     = os.environ["TELEGRAM_BOT_TOKEN"]
ANTHROPIC_KEY = os.environ["ANTHROPIC_API_KEY"]
DB_PATH       = os.environ.get("DB_PATH", "facturas.db")

ALLOWED_USERS_ENV = os.environ.get("ALLOWED_USERS", "")
ALLOWED_USERS = set(
    int(x.strip()) for x in ALLOWED_USERS_ENV.split(",") if x.strip().isdigit()
)

# Contraseña para iniciar el bot. Se puede cambiar con la variable BOT_PASSWORD
# en Railway; por defecto es la acordada con el cliente.
BOT_PASSWORD = os.environ.get("BOT_PASSWORD", "/Juan2202")

# RNC propio (el receptor/comprador). Se excluye del campo "rnc" que siempre
# debe ser el del EMISOR de la factura.
CONSUMER_RNC = os.environ.get("CONSUMER_RNC", "131545157")

# Conversation states  (simplificado: foto → ubicación → categoría → confirmar)
(
    S_PHOTO,
    S_LOCATION,
    S_CATEGORY,
    S_CONFIRM,
    S_EDIT_SELECT,
    S_EDIT_VALUE,
) = range(6)

LOCATIONS   = ["Punta Cana", "Santo Domingo"]
CATEGORIES  = ["Casa", "Obra"]

# Botones del menú fijo (abajo del chat). El texto enviado por cada botón se
# enruta al comando equivalente, así no hay que escribir comandos a mano.
BTN_NUEVA      = "🧾 Nueva factura"
BTN_RESUMEN    = "📊 Resumen"
BTN_EXPORTAR   = "📥 Descargar Excel"
BTN_PENDIENTES = "⚠️ Pendientes"
BTN_AYUDA      = "❓ Ayuda"

def main_menu_keyboard() -> ReplyKeyboardMarkup:
    """Menú permanente de botones grandes; evita tener que escribir comandos."""
    return ReplyKeyboardMarkup(
        [
            [KeyboardButton(BTN_NUEVA)],
            [KeyboardButton(BTN_RESUMEN), KeyboardButton(BTN_EXPORTAR)],
            [KeyboardButton(BTN_PENDIENTES), KeyboardButton(BTN_AYUDA)],
        ],
        resize_keyboard=True,
        is_persistent=True,
        input_field_placeholder="Toca un botón o envía una foto de la factura",
    )

EDITABLE_FIELDS = {
    "total":      "💰 Total",
    "itbis":      "🧾 ITBIS",
    "base":       "📦 Base",
    "propina":    "🍽️ Propina",
    "ncf":        "🔢 NCF",
    "rnc":        "🏢 RNC",
    "nombre":     "🏪 Nombre proveedor",
    "fecha":      "📅 Fecha",
    "metodo":     "💳 Método pago",
    "obs":        "📝 Observaciones",
    "tipo_gasto": "🏷️ Tipo de gasto",
}

METODO_LABELS = {
    "EFECTIVO":       "💵 Efectivo",
    "TARJETA_CREDITO":"💳 T.Crédito",
    "TARJETA_DEBITO": "💳 T.Débito",
    "TRANSFERENCIA":  "🏦 Transfer.",
    "CHEQUE":         "📄 Cheque",
    "CREDITO":        "📋 Crédito",
}

# Tipos de gasto DGII Formato 606 (Art. 4 del Reglamento 293-11)
TIPO_GASTO_OPTIONS = [
    ("01", "Gastos de personal"),
    ("02", "Trabajos, suministros y servicios"),
    ("03", "Arrendamientos"),
    ("04", "Gastos de activos fijos"),
    ("05", "Gastos de representación"),
    ("06", "Otras deducciones admitidas"),
    ("07", "Gastos financieros"),
    ("08", "Gastos extraordinarios"),
    ("09", "Compras para costo de ventas"),
    ("10", "Adquisición de activos depreciables"),
    ("11", "Seguros"),
]
TIPO_GASTO_DICT = {k: v for k, v in TIPO_GASTO_OPTIONS}

# ──────────────────────────────────────────────────────────────
# DATABASE
# ──────────────────────────────────────────────────────────────

def get_db() -> sqlite3.Connection:
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def init_db():
    with get_db() as conn:
        conn.execute("""
            CREATE TABLE IF NOT EXISTS facturas (
                id              INTEGER PRIMARY KEY AUTOINCREMENT,
                mes             TEXT NOT NULL,
                location        TEXT,
                category        TEXT,
                filename        TEXT,
                rnc             TEXT,
                ncf             TEXT,
                nombre          TEXT,
                fecha_comp      TEXT,
                fecha_pago      TEXT,
                total           REAL DEFAULT 0,
                itbis           REAL DEFAULT 0,
                base            REAL DEFAULT 0,
                propina         REAL DEFAULT 0,
                metodo          TEXT,
                tipo_cf         TEXT,
                observaciones   TEXT,
                qr_verified     INTEGER DEFAULT 0,
                nivel_confianza TEXT,
                advertencias    TEXT,
                needs_review    INTEGER DEFAULT 0,
                raw_json        TEXT,
                usuario         TEXT,
                created_at      TEXT DEFAULT (datetime('now'))
            )
        """)
        # Migraciones: agregar columnas nuevas si la tabla ya existía.
        cols = {r[1] for r in conn.execute("PRAGMA table_info(facturas)")}
        if "usuario" not in cols:
            conn.execute("ALTER TABLE facturas ADD COLUMN usuario TEXT")
        if "tipo_gasto" not in cols:
            conn.execute("ALTER TABLE facturas ADD COLUMN tipo_gasto TEXT DEFAULT '02'")
        conn.commit()


def save_factura(mes: str, location: str, category: str, data: dict,
                 usuario: str = "", tipo_gasto: str = "02") -> int:
    with get_db() as conn:
        cur = conn.execute("""
            INSERT INTO facturas
              (mes, location, category, filename, rnc, ncf, nombre,
               fecha_comp, fecha_pago, total, itbis, base, propina,
               metodo, tipo_cf, observaciones, qr_verified,
               nivel_confianza, advertencias, needs_review, raw_json, usuario, tipo_gasto)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        """, (
            mes, location, category,
            data.get("_filename", ""),
            data.get("rnc", ""),
            data.get("ncf", ""),
            data.get("nombre_proveedor", ""),
            data.get("fecha_comprobante", ""),
            data.get("fecha_pago", ""),
            float(data.get("total_facturado") or 0),
            float(data.get("itbis") or 0),
            float(data.get("monto_sin_itbis") or 0),
            float(data.get("propina") or 0),
            data.get("metodo_pago", ""),
            data.get("tipo_comprobante", ""),
            data.get("observaciones", ""),
            1 if data.get("_qr_verified") else 0,
            data.get("nivel_confianza", "ALTO"),
            json.dumps(data.get("_warnings") or []),
            1 if data.get("_needs_review") else 0,
            json.dumps(data),
            usuario,
            tipo_gasto,
        ))
        conn.commit()
        return cur.lastrowid


def get_facturas(mes: str) -> list[dict]:
    with get_db() as conn:
        rows = conn.execute(
            "SELECT * FROM facturas WHERE mes=? ORDER BY id", (mes,)
        ).fetchall()
    return [dict(r) for r in rows]


def meses_con_facturas() -> list[str]:
    """Meses (YYYY-MM) que ya tienen facturas guardadas, del más reciente al más viejo."""
    with get_db() as conn:
        rows = conn.execute(
            "SELECT mes, COUNT(*) n FROM facturas GROUP BY mes ORDER BY mes DESC"
        ).fetchall()
    return [(r[0], r[1]) for r in rows]


def check_duplicate_ncf(ncf: str) -> dict | None:
    """Busca si el NCF ya existe en cualquier mes. Devuelve la fila o None."""
    if not ncf:
        return None
    with get_db() as conn:
        row = conn.execute(
            "SELECT * FROM facturas WHERE ncf=? LIMIT 1", (ncf,)
        ).fetchone()
    return dict(row) if row else None


def delete_last(mes: str) -> dict | None:
    with get_db() as conn:
        row = conn.execute(
            "SELECT * FROM facturas WHERE mes=? ORDER BY id DESC LIMIT 1", (mes,)
        ).fetchone()
        if row:
            conn.execute("DELETE FROM facturas WHERE id=?", (row["id"],))
            conn.commit()
            return dict(row)
    return None


# ──────────────────────────────────────────────────────────────
# EXTRACCIÓN (QR + Claude Vision)
# ──────────────────────────────────────────────────────────────

def _build_extraction_prompt() -> str:
    consumer_line = (
        f"\n⚠️ RNC DEL COMPRADOR (RECEPTOR) A IGNORAR: {CONSUMER_RNC}\n"
        f"   Este RNC aparece en la factura como receptor/consumidor — NUNCA lo uses como 'rnc'.\n"
        f"   El campo 'rnc' es SIEMPRE el RNC del EMISOR (quien emite la factura, el proveedor)."
        if CONSUMER_RNC else ""
    )
    return f"""Eres un experto en facturación dominicana y fiscalización DGII.

━━━ REGLA DE ORO SOBRE RNC ━━━
'rnc' = RNC del EMISOR (el negocio que emite y firma el comprobante fiscal).
NUNCA uses el RNC del receptor/comprador.{consumer_line}

━━━ REGLA DE ORO SOBRE MONTOS ━━━
"total_facturado" = importe FINAL pagado (el número más grande al fondo del ticket).
"itbis"           = monto del ITBIS ya incluido en ese total.
"monto_sin_itbis" = total_facturado − itbis − propina.
NUNCA sumes ITBIS al total — ya está dentro del total.

EJEMPLOS:
• SubTotal 750.00 | ITBIS 135.00 | Propina 75.00 | TOTAL 960.00
  → total=960, itbis=135, propina=75, base=750 ✓
• Total General 6,817.08 | ITBIS 16%=119.85 | ITBIS 18%=246.08
  → total=6817.08, itbis=365.93, base=6451.15 ✓
• Farmacia sin ITBIS: Total 904.80
  → total=904.80, itbis=0, base=904.80 ✓

ITBIS: suma TODAS las tasas (16%+18%) en un solo número.
ISC (seguros/licores): NO es ITBIS — ponlo en observaciones.
PROPINA: 10% solo en restaurantes/bares.
PAGOS PARCIALES: si la cuenta se pagó en varias tarjetas/vouchers (ej. voucher
CARDNET por la mitad), NO importa: usa SIEMPRE el TOTAL COMPLETO de la factura
fiscal. Los vouchers de pago parcial no son facturas separadas.
MÉTODO: EFECTIVO | TARJETA_CREDITO | TARJETA_DEBITO | TRANSFERENCIA | CHEQUE | CREDITO
NCF papel: B+2+8 dígitos=11 chars. NCF electrónico: E+2+10 dígitos=13 chars.
NCF SIN guiones ni espacios: 'B01-0001234' → 'B010001234'. Solo letras y números.
TIPO: B01/E31=CREDITO_FISCAL | B02/E32=CONSUMIDOR_FINAL | B14/E34=REGIMEN_ESPECIAL

nivel_confianza: ALTO|MEDIO|BAJO

Responde SOLO con JSON válido, sin markdown:
{{"rnc":"","ncf":"","fecha_comprobante":"YYYY-MM-DD","fecha_pago":"YYYY-MM-DD",
"total_facturado":0,"itbis":0,"monto_sin_itbis":0,"propina":0,
"metodo_pago":"","nombre_proveedor":"","tipo_comprobante":"","observaciones":null,
"nivel_confianza":"ALTO"}}"""

EXTRACTION_PROMPT: str = ""  # populated in main() after CONSUMER_RNC is set


def decode_qr(image_bytes: bytes) -> str | None:
    try:
        import numpy as np
        from PIL import Image
        from pyzbar import pyzbar
        img = Image.open(io.BytesIO(image_bytes))
        for d in pyzbar.decode(np.array(img)):
            return d.data.decode("utf-8", errors="ignore")
    except Exception:
        pass
    try:
        import cv2
        import numpy as np
        from PIL import Image
        img_np = np.array(Image.open(io.BytesIO(image_bytes)).convert("RGB"))
        img_bgr = cv2.cvtColor(img_np, cv2.COLOR_RGB2BGR)
        det = cv2.QRCodeDetector()
        for scale in [1.0, 2.0, 3.0]:
            h, w = img_bgr.shape[:2]
            r = cv2.resize(img_bgr, (int(w*scale), int(h*scale)), interpolation=cv2.INTER_CUBIC)
            data, _, _ = det.detectAndDecode(r)
            if data:
                return data
    except Exception:
        pass
    return None


def parse_ecf_url(qr_text: str) -> dict | None:
    if not qr_text or "ConsultaTimbre" not in qr_text:
        return None
    try:
        params = parse_qs(urlparse(qr_text).query)
        def get(k):
            for key, v in params.items():
                if key.lower() == k.lower(): return v[0]
            return None
        rnc  = re.sub(r"\D", "", get("RNCEmisor") or get("RncEmisor") or "")
        ncf  = clean_ncf(get("ENCF"))
        tot  = float((get("MontoTotal") or "0").replace(",", "."))
        fecha = None
        if fe := get("FechaEmision"):
            try:
                d, m, y = fe.split("-"); fecha = f"{y}-{m}-{d}"
            except Exception:
                pass
        if rnc and ncf and tot:
            return {"rnc": rnc, "ncf": ncf, "total": tot, "fecha": fecha}
    except Exception:
        pass
    return None


def validate_and_fix(data: dict, qr: dict | None = None) -> dict:
    if data.get("_error"):
        data["_needs_review"] = True
        data["_warnings"] = ["Error de extracción"]
        return data

    if qr:
        if qr.get("rnc"):   data["rnc"] = qr["rnc"]
        if qr.get("ncf"):   data["ncf"] = qr["ncf"]
        if qr.get("total"): data["total_facturado"] = qr["total"]
        if qr.get("fecha"): data.setdefault("fecha_comprobante", qr["fecha"])
        data["_qr_verified"] = True

    def f(v):
        try: return float(v) if v not in (None, "", "null") else 0.0
        except: return 0.0

    total   = f(data.get("total_facturado"))
    itbis   = f(data.get("itbis"))
    propina = f(data.get("propina"))
    base    = f(data.get("monto_sin_itbis"))

    if total == 0 and (base + itbis) > 0:
        total = base + itbis + propina
    if base == 0 and total > 0:
        base = round(total - itbis - propina, 2)

    data["total_facturado"]  = total
    data["itbis"]            = itbis
    data["propina"]          = propina
    data["monto_sin_itbis"]  = base

    warns = []
    calc = round(base + itbis + propina, 2)
    if abs(calc - total) > 1.0:
        warns.append(f"⚠ Descuadre: {base:.2f}+{itbis:.2f}+{propina:.2f}={calc:.2f} ≠ {total:.2f}")

    rnc = re.sub(r"\D", "", str(data.get("rnc") or ""))
    if len(rnc) not in (9, 11):
        warns.append(f"⚠ RNC '{rnc}' tiene {len(rnc)} dígitos (esperado 9 u 11)")

    ncf = clean_ncf(data.get("ncf"))
    data["ncf"] = ncf  # guardar sin guiones ni espacios
    if ncf:
        expected = 11 if ncf.startswith("B") else 13 if ncf.startswith("E") else None
        if expected and len(ncf) != expected:
            warns.append(f"⚠ NCF tiene {len(ncf)} chars (esperado {expected}: {ncf})")

    if (data.get("nivel_confianza") or "").upper() == "BAJO":
        warns.append("⚠ Imagen poco legible — verificar")

    data["_warnings"]     = warns
    data["_needs_review"] = len(warns) > 0
    return data


async def extract_invoice(image_bytes: bytes, filename: str) -> dict:
    qr_raw = decode_qr(image_bytes)
    qr     = parse_ecf_url(qr_raw) if qr_raw else None

    hint = ""
    if qr:
        hint = (f"\n\n⚡ QR verificado — usar exactamente:\n"
                f"RNC={qr['rnc']} | NCF={qr['ncf']} | Total={qr['total']}")
        if qr.get("fecha"): hint += f" | Fecha={qr['fecha']}"

    img_b64 = base64.standard_b64encode(image_bytes).decode()
    client  = anthropic.AsyncAnthropic(api_key=ANTHROPIC_KEY)

    try:
        resp = await client.messages.create(
            model="claude-opus-4-8",
            max_tokens=1024,
            system=[{"type": "text", "text": _build_extraction_prompt(),
                     "cache_control": {"type": "ephemeral"}}],
            messages=[{"role": "user", "content": [
                {"type": "image", "source": {
                    "type": "base64", "media_type": "image/jpeg", "data": img_b64}},
                {"type": "text",
                 "text": f"Extrae datos de esta factura ({filename}).{hint}\nResponde solo JSON."},
            ]}],
        )
        text = resp.content[0].text.strip()
        if text.startswith("```"):
            text = "\n".join(text.split("\n")[1:]).removesuffix("```").strip()
        data = json.loads(text)
        data["_filename"] = filename
        data["_error"]    = None
    except json.JSONDecodeError as e:
        data = {"_filename": filename, "_error": f"JSON inválido: {e}"}
    except Exception as e:
        data = {"_filename": filename, "_error": str(e)}

    return validate_and_fix(data, qr)


# ──────────────────────────────────────────────────────────────
# FORMATO DE MENSAJES
# ──────────────────────────────────────────────────────────────

def format_review_message(data: dict, location: str, category: str,
                          mes: str | None = None,
                          tipo_gasto: str | None = None) -> str:
    """Mensaje de revisión completo, listo para mostrar con confirm_keyboard."""
    qr_badge  = "✅ QR verificado" if data.get("_qr_verified") else "🤖 Extraído por IA"
    conf      = data.get("nivel_confianza", "ALTO")
    conf_icon = "🟢" if conf == "ALTO" else "🟡" if conf == "MEDIO" else "🔴"
    warns     = data.get("_warnings") or []
    metodo    = METODO_LABELS.get(data.get("metodo_pago", ""), data.get("metodo_pago", "—"))
    destino   = mes or mes_from_fecha(data)

    msg = (
        f"🧾 *Revisión de factura*\n"
        f"📍 {location}  •  🏷️ {category}\n"
        f"{qr_badge}  {conf_icon} Confianza: {conf}\n"
        f"{'─'*30}\n"
        f"🏪 *{data.get('nombre_proveedor') or '—'}*\n"
        f"🔢 NCF: `{data.get('ncf') or '—'}`\n"
        f"🏢 RNC: `{data.get('rnc') or '—'}`\n"
        f"📅 Fecha: {data.get('fecha_comprobante') or '—'}\n"
        f"{'─'*30}\n"
        f"📦 Base sin ITBIS:  RD$ *{float(data.get('monto_sin_itbis') or 0):,.2f}*\n"
        f"🧾 ITBIS:           RD$ *{float(data.get('itbis') or 0):,.2f}*\n"
    )
    if float(data.get("propina") or 0) > 0:
        msg += f"🍽️ Propina 10%:     RD$ *{float(data.get('propina') or 0):,.2f}*\n"
    msg += (
        f"💰 *TOTAL:          RD$ {float(data.get('total_facturado') or 0):,.2f}*\n"
        f"{'─'*30}\n"
        f"{metodo}\n"
    )
    if tipo_gasto:
        tg_label = TIPO_GASTO_DICT.get(tipo_gasto, tipo_gasto)
        msg += f"🏷️ Tipo de gasto: *{tipo_gasto}* — {tg_label}\n"
    if data.get("observaciones"):
        msg += f"📝 {data['observaciones']}\n"
    if warns:
        msg += f"\n{'─'*30}\n" + "\n".join(warns)
    return msg


MES_NOMBRES = {
    "01": "Ene", "02": "Feb", "03": "Mar", "04": "Abr", "05": "May", "06": "Jun",
    "07": "Jul", "08": "Ago", "09": "Sep", "10": "Oct", "11": "Nov", "12": "Dic",
}

def mes_label(ym: str) -> str:
    """'2026-05' → 'May 2026' para mostrar en botones."""
    try:
        y, m = ym.split("-")
        return f"{MES_NOMBRES.get(m, m)} {y}"
    except Exception:
        return ym

def month_keyboard() -> InlineKeyboardMarkup:
    """Botones para elegir a qué Excel (mes) va la factura."""
    rows, row = [], []
    for ym in month_options():
        row.append(InlineKeyboardButton(f"📅 {mes_label(ym)}", callback_data=f"mes_{ym}"))
        if len(row) == 3:
            rows.append(row); row = []
    if row:
        rows.append(row)
    rows.append([InlineKeyboardButton("✨ Auto (por fecha de la factura)",
                                      callback_data="mes_AUTO")])
    return InlineKeyboardMarkup(rows)


def mes_picker(prefix: str) -> InlineKeyboardMarkup | None:
    """Botones con los meses que ya tienen facturas, para /resumen y /exportar.
    `prefix` distingue la acción (p.ej. 'res' o 'exp'). Devuelve None si no hay datos."""
    meses = meses_con_facturas()
    if not meses:
        return None
    rows, row = [], []
    for ym, n in meses:
        row.append(InlineKeyboardButton(
            f"📅 {mes_label(ym)} ({n})", callback_data=f"{prefix}_{ym}"))
        if len(row) == 2:
            rows.append(row); row = []
    if row:
        rows.append(row)
    return InlineKeyboardMarkup(rows)


def suggest_tipo_gasto(data: dict) -> str:
    """Auto-detecta el tipo de gasto DGII más probable según los datos de la factura."""
    propina = float(data.get("propina") or 0)
    if propina > 0:
        return "05"  # Restaurantes/bares → Gastos de representación
    ncf = (data.get("ncf") or "").upper()
    if ncf.startswith(("B11", "E41")):
        return "07"  # Nota de crédito → probablemente financiero
    nombre = (data.get("nombre_proveedor") or "").upper()
    if any(x in nombre for x in ("BANCO", "FINANCIER", "PRESTAMO", "PRÉSTAMO", "CREDITO", "CRÉDITO")):
        return "07"  # Gastos financieros
    if any(x in nombre for x in ("SEGURO", "SEGUROS", "INSURANCE")):
        return "11"  # Seguros
    if any(x in nombre for x in ("ALQUILER", "ARRENDAMIENTO", "RENTA LOCAL")):
        return "03"  # Arrendamientos
    obs = (data.get("observaciones") or "").upper()
    if any(x in obs for x in ("NOMINA", "NÓMINA", "SALARIO", "PERSONAL")):
        return "01"  # Gastos de personal
    return "02"  # Default: trabajos, suministros y servicios


def tipo_gasto_keyboard(suggested: str = "") -> InlineKeyboardMarkup:
    rows = []
    for code, label in TIPO_GASTO_OPTIONS:
        star = "⭐ " if code == suggested else ""
        rows.append([InlineKeyboardButton(
            f"{star}{code} – {label}", callback_data=f"tg_{code}"
        )])
    rows.append([InlineKeyboardButton("← Cambiar categoría", callback_data="back_to_category")])
    return InlineKeyboardMarkup(rows)


def confirm_keyboard(mes: str = "") -> InlineKeyboardMarkup:
    mes_txt = f"📅 {mes_label(mes)} — cambiar mes" if mes else "📅 Cambiar mes"
    return InlineKeyboardMarkup([
        [
            InlineKeyboardButton("✅ Aceptar",  callback_data="confirm_accept"),
            InlineKeyboardButton("✏️ Corregir", callback_data="confirm_edit"),
            InlineKeyboardButton("❌ Cancelar", callback_data="confirm_cancel"),
        ],
        [InlineKeyboardButton(mes_txt, callback_data="confirm_change_mes")],
    ])


def tipo_gasto_edit_keyboard(selected: str = "") -> InlineKeyboardMarkup:
    rows = []
    for code, label in TIPO_GASTO_OPTIONS:
        mark = "✅ " if code == selected else ""
        rows.append([InlineKeyboardButton(
            f"{mark}{code} – {label}", callback_data=f"edittg_{code}"
        )])
    rows.append([InlineKeyboardButton("↩️ Volver a revisión", callback_data="edittg_back")])
    return InlineKeyboardMarkup(rows)


def edit_keyboard() -> InlineKeyboardMarkup:
    """Keyboard for selecting which field to edit."""
    buttons = []
    row = []
    for key, label in EDITABLE_FIELDS.items():
        row.append(InlineKeyboardButton(label, callback_data=f"edit_{key}"))
        if len(row) == 2:
            buttons.append(row)
            row = []
    if row:
        buttons.append(row)
    buttons.append([InlineKeyboardButton("↩️ Volver a revisión", callback_data="edit_back")])
    return InlineKeyboardMarkup(buttons)


# ──────────────────────────────────────────────────────────────
# HELPERS
# ──────────────────────────────────────────────────────────────

def current_mes() -> str:
    return datetime.now().strftime("%Y-%m")

def get_mes(context) -> str:
    return context.user_data.get("mes_activo", current_mes())

def clean_ncf(s: str) -> str:
    """Normaliza el NCF: mayúsculas y solo letras/números (quita guiones, espacios).
    Ej.: 'B01-0001234' → 'B010001234'  •  'E31-0000012345' → 'E310000012345'."""
    return re.sub(r"[^A-Z0-9]", "", (s or "").upper())

def month_options() -> list[str]:
    """Meses sugeridos para los botones: los 3 anteriores, el actual y el siguiente."""
    now = datetime.now()
    y, m = now.year, now.month
    opts = []
    for delta in range(-3, 2):  # -3 .. +1
        mm = m + delta
        yy = y + (mm - 1) // 12
        mm = (mm - 1) % 12 + 1
        opts.append(f"{yy:04d}-{mm:02d}")
    return opts

def mes_from_fecha(data: dict) -> str:
    """
    Mes (YYYY-MM) al que pertenece la factura, según su fecha de comprobante.
    Así cada factura cae sola en el Excel de su mes (mayo→junio→julio…).
    Si la fecha no es legible, usa el mes actual del calendario.
    """
    fecha = str(data.get("fecha_comprobante") or "")
    m = re.match(r"^(\d{4})-(\d{2})", fecha)
    if m:
        return f"{m.group(1)}-{m.group(2)}"
    return current_mes()

def is_allowed(update: Update) -> bool:
    if not ALLOWED_USERS: return True
    return update.effective_user.id in ALLOWED_USERS

def resolve_mes(context, data: dict) -> str:
    """Mes destino de la factura: el elegido con los botones, o automático
    (según la fecha del comprobante) si el usuario eligió 'Auto'."""
    elegido = context.user_data.get("mes_elegido")
    if elegido and elegido != "AUTO":
        return elegido
    return mes_from_fecha(data)

def user_label(update: Update) -> str:
    """Nombre legible de quien subió la factura (para registrar en el Excel)."""
    u = update.effective_user
    if not u:
        return ""
    if u.username:
        return f"@{u.username}"
    nombre = " ".join(filter(None, [u.first_name, u.last_name])).strip()
    return nombre or str(u.id)


def _CLEANUP(context):
    """Limpia los datos de la factura activa del contexto."""
    for key in ("pending_invoice", "location", "category", "tipo_gasto",
                "edit_field", "pending_photo_id", "mes_elegido"):
        context.user_data.pop(key, None)


def _review_msg_and_kb(context) -> tuple[str, InlineKeyboardMarkup]:
    """Genera el mensaje de revisión y el teclado de confirmación actuales."""
    data     = context.user_data.get("pending_invoice", {})
    location = context.user_data.get("location", "—")
    category = context.user_data.get("category", "—")
    tg       = context.user_data.get("tipo_gasto", "02")
    mes      = resolve_mes(context, data)
    msg = format_review_message(data, location, category, mes, tg)
    return msg, confirm_keyboard(mes)


# ──────────────────────────────────────────────────────────────
# CONVERSATION: FLUJO DE CAPTURA
# Flujo: foto → ubicación → categoría → revisión+confirmar
# ──────────────────────────────────────────────────────────────

async def start_flow(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Inicio del flujo. Si llegó foto directamente, la procesa ya."""
    if not is_allowed(update): return ConversationHandler.END
    _CLEANUP(context)
    if update.message and update.message.photo:
        return await _process_photo(update.message, context, update.message.photo[-1].file_id)
    await update.message.reply_text("📸 Envía la foto de la factura.")
    return S_PHOTO


async def _process_photo(reply_to, context: ContextTypes.DEFAULT_TYPE, file_id: str) -> int:
    """Descarga la foto, extrae datos con IA, verifica duplicados y pregunta la ubicación."""
    proc = await reply_to.reply_text("⏳ Analizando factura con IA...")
    file = await context.bot.get_file(file_id)
    buf  = io.BytesIO()
    await file.download_to_memory(buf)
    data = await extract_invoice(buf.getvalue(), f"{file_id}.jpg")
    await proc.delete()

    if data.get("_error"):
        await reply_to.reply_text(
            f"❌ No pude leer la factura: {data['_error']}\n\n"
            "Intenta con mejor iluminación o más cerca.",
        )
        return ConversationHandler.END

    ncf = data.get("ncf", "")
    if ncf:
        dup = check_duplicate_ncf(ncf)
        if dup:
            await reply_to.reply_text(
                f"⚠️ *Factura duplicada — no procesada*\n\n"
                f"El NCF `{ncf}` ya existe en el sistema:\n"
                f"🏪 {dup.get('nombre') or '?'}  •  RD$ {dup.get('total', 0):,.2f}  ({dup.get('mes', '')})\n\n"
                f"Envía otra foto o usa el menú.",
                parse_mode="Markdown",
            )
            return S_PHOTO

    context.user_data["pending_invoice"] = data
    context.user_data["tipo_gasto"] = suggest_tipo_gasto(data)

    nombre = data.get("nombre_proveedor") or "Factura leída"
    total  = float(data.get("total_facturado") or 0)
    kb = InlineKeyboardMarkup([[
        InlineKeyboardButton(f"📍 {loc}", callback_data=f"loc_{loc}") for loc in LOCATIONS
    ]])
    await reply_to.reply_text(
        f"✅ *{nombre}* — RD$ {total:,.2f}\n\n📍 ¿Dónde fue esta compra?",
        reply_markup=kb,
        parse_mode="Markdown",
    )
    return S_LOCATION


async def got_photo(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Foto recibida en estado S_PHOTO."""
    if not is_allowed(update): return ConversationHandler.END
    return await _process_photo(update.message, context, update.message.photo[-1].file_id)


async def got_location(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Ubicación elegida → preguntar categoría."""
    query = update.callback_query
    await query.answer()
    location = query.data.replace("loc_", "")
    context.user_data["location"] = location
    kb = InlineKeyboardMarkup([
        [InlineKeyboardButton(f"🏷️ {cat}", callback_data=f"cat_{cat}") for cat in CATEGORIES],
        [InlineKeyboardButton("← Cambiar ubicación", callback_data="back_to_location")],
    ])
    await query.edit_message_text(
        f"📍 *{location}*\n\n🏷️ ¿Para qué es la compra?",
        reply_markup=kb,
        parse_mode="Markdown",
    )
    return S_CATEGORY


async def back_to_location(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """← Volver a la pregunta de ubicación."""
    query = update.callback_query
    await query.answer()
    data   = context.user_data.get("pending_invoice", {})
    nombre = data.get("nombre_proveedor") or "Factura leída"
    total  = float(data.get("total_facturado") or 0)
    kb = InlineKeyboardMarkup([[
        InlineKeyboardButton(f"📍 {loc}", callback_data=f"loc_{loc}") for loc in LOCATIONS
    ]])
    await query.edit_message_text(
        f"✅ *{nombre}* — RD$ {total:,.2f}\n\n📍 ¿Dónde fue esta compra?",
        reply_markup=kb,
        parse_mode="Markdown",
    )
    return S_LOCATION


async def got_category(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Categoría elegida → mostrar revisión completa con tipo de gasto auto-detectado."""
    query = update.callback_query
    await query.answer()
    context.user_data["category"] = query.data.replace("cat_", "")
    msg, kb = _review_msg_and_kb(context)
    await query.edit_message_text(msg, reply_markup=kb, parse_mode="Markdown")
    return S_CONFIRM


async def on_confirm_change_mes(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Botón 'Cambiar mes' en la revisión → mostrar selector de mes."""
    query = update.callback_query
    await query.answer()
    rows, row = [], []
    for ym in month_options():
        row.append(InlineKeyboardButton(f"📅 {mes_label(ym)}", callback_data=f"cmes_{ym}"))
        if len(row) == 3:
            rows.append(row); row = []
    if row:
        rows.append(row)
    rows.append([InlineKeyboardButton("✨ Auto (por fecha de la factura)", callback_data="cmes_AUTO")])
    rows.append([InlineKeyboardButton("← Volver a revisión", callback_data="cmes_back")])
    await query.edit_message_text(
        "📅 ¿A qué mes quieres asignar esta factura?",
        reply_markup=InlineKeyboardMarkup(rows),
    )
    return S_CONFIRM


async def on_cmes_picked(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Mes elegido desde la revisión → volver a revisión con mes actualizado."""
    query = update.callback_query
    await query.answer()
    val = query.data.replace("cmes_", "")
    if val != "back":
        context.user_data["mes_elegido"] = val
    msg, kb = _review_msg_and_kb(context)
    await query.edit_message_text(msg, reply_markup=kb, parse_mode="Markdown")
    return S_CONFIRM


async def confirm_accept(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Guardar la factura."""
    query = update.callback_query
    await query.answer()

    data       = context.user_data.get("pending_invoice", {})
    location   = context.user_data.get("location", "—")
    category   = context.user_data.get("category", "—")
    mes        = resolve_mes(context, data)
    tipo_gasto = context.user_data.get("tipo_gasto", "02")

    context.user_data["mes_activo"] = mes
    fac_id = save_factura(mes, location, category, data, user_label(update), tipo_gasto)

    facturas  = get_facturas(mes)
    total_mes = sum(f["total"] for f in facturas)

    drive_line = ""
    if drive_sync.is_configured():
        try:
            xlsx = build_excel(facturas, mes)
            link = await asyncio.to_thread(drive_sync.sync_excel, xlsx, f"606_{mes}.xlsx")
            drive_line = f"\n☁️ [Excel actualizado en Drive]({link})\n" if link else "\n⚠️ No se pudo subir a Drive.\n"
        except Exception as e:
            log.error("Drive sync error: %s", e)
            drive_line = "\n⚠️ Error al subir a Drive.\n"

    tg_label  = TIPO_GASTO_DICT.get(tipo_gasto, tipo_gasto)
    saved_msg = (
        f"✅ *Factura #{fac_id} guardada*\n\n"
        f"🏪 {data.get('nombre_proveedor','?')}\n"
        f"💰 RD$ {float(data.get('total_facturado',0)):,.2f}\n"
        f"🏷️ {tipo_gasto} — {tg_label}\n"
        f"📊 *{mes}* — {len(facturas)} facturas  •  RD$ {total_mes:,.2f}"
        f"{drive_line}"
    )
    kb = InlineKeyboardMarkup([[
        InlineKeyboardButton("📸 Subir otra factura", callback_data="otra_factura"),
        InlineKeyboardButton("✅ Terminar",           callback_data="fin_lote"),
    ]])
    await query.edit_message_text(
        saved_msg, reply_markup=kb, parse_mode="Markdown", disable_web_page_preview=True
    )
    _CLEANUP(context)
    return S_PHOTO


async def otra_factura(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Callback 'Subir otra' → quedarse en S_PHOTO esperando la siguiente foto."""
    query = update.callback_query
    await query.answer("📸 Envía la siguiente foto")
    await query.edit_message_text(
        query.message.text + "\n\n📸 Listo, envía la siguiente factura.",
        parse_mode="Markdown",
        disable_web_page_preview=True,
    )
    return S_PHOTO


async def fin_lote(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Callback 'Terminar' → cerrar conversación."""
    query = update.callback_query
    await query.answer()
    await query.edit_message_text(
        query.message.text + "\n\n_Usa Resumen o Descargar Excel para ver los datos._",
        parse_mode="Markdown",
        disable_web_page_preview=True,
    )
    return ConversationHandler.END


async def confirm_edit(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Mostrar selector de campos para editar."""
    query = update.callback_query
    await query.answer()
    await query.edit_message_text(
        "✏️ *¿Qué campo deseas corregir?*\n\nSelecciona el campo:",
        reply_markup=edit_keyboard(),
        parse_mode="Markdown",
    )
    return S_EDIT_SELECT


async def edit_field_selected(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Campo seleccionado → pedir nuevo valor (o mostrar selector si es tipo_gasto)."""
    query = update.callback_query
    await query.answer()

    if query.data == "edit_back":
        msg, kb = _review_msg_and_kb(context)
        await query.edit_message_text(msg, reply_markup=kb, parse_mode="Markdown")
        return S_CONFIRM

    field = query.data.replace("edit_", "")
    context.user_data["edit_field"] = field

    if field == "tipo_gasto":
        current_tg = context.user_data.get("tipo_gasto", "02")
        await query.edit_message_text(
            "🏷️ *Selecciona el tipo de gasto DGII:*",
            reply_markup=tipo_gasto_edit_keyboard(current_tg),
            parse_mode="Markdown",
        )
        return S_EDIT_SELECT

    data  = context.user_data.get("pending_invoice", {})
    label = EDITABLE_FIELDS.get(field, field)

    current_values = {
        "total":   f"RD$ {float(data.get('total_facturado',0)):,.2f}",
        "itbis":   f"RD$ {float(data.get('itbis',0)):,.2f}",
        "base":    f"RD$ {float(data.get('monto_sin_itbis',0)):,.2f}",
        "propina": f"RD$ {float(data.get('propina',0)):,.2f}",
        "ncf":     data.get("ncf",""),
        "rnc":     data.get("rnc",""),
        "nombre":  data.get("nombre_proveedor",""),
        "fecha":   data.get("fecha_comprobante",""),
        "metodo":  data.get("metodo_pago",""),
        "obs":     data.get("observaciones",""),
    }

    metodo_hint = ""
    if field == "metodo":
        metodo_hint = ("\n\nOpciones válidas:\n"
                       "EFECTIVO | TARJETA_CREDITO | TARJETA_DEBITO\n"
                       "TRANSFERENCIA | CHEQUE | CREDITO")

    await query.edit_message_text(
        f"✏️ *Corregir: {label}*\n\n"
        f"Valor actual: `{current_values.get(field,'—')}`\n\n"
        f"Escribe el nuevo valor:{metodo_hint}",
        parse_mode="Markdown",
    )
    return S_EDIT_VALUE


async def edit_tipo_gasto_pick(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Tipo de gasto elegido desde la edición → volver a revisión."""
    query = update.callback_query
    await query.answer()
    val = query.data.replace("edittg_", "")
    if val != "back":
        context.user_data["tipo_gasto"] = val
    msg, kb = _review_msg_and_kb(context)
    await query.edit_message_text(msg, reply_markup=kb, parse_mode="Markdown")
    return S_CONFIRM


async def edit_value_received(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Aplica el nuevo valor y vuelve a la revisión."""
    if not is_allowed(update): return ConversationHandler.END

    field    = context.user_data.get("edit_field")
    new_val  = update.message.text.strip()
    data     = context.user_data.get("pending_invoice", {})

    try:
        if field == "total":
            data["total_facturado"] = float(new_val.replace(",", ".").replace("RD$", "").strip())
        elif field == "itbis":
            data["itbis"] = float(new_val.replace(",", ".").replace("RD$", "").strip())
        elif field == "base":
            data["monto_sin_itbis"] = float(new_val.replace(",", ".").replace("RD$", "").strip())
        elif field == "propina":
            data["propina"] = float(new_val.replace(",", ".").replace("RD$", "").strip())
        elif field == "ncf":
            data["ncf"] = clean_ncf(new_val)
        elif field == "rnc":
            data["rnc"] = re.sub(r"\D", "", new_val)
        elif field == "nombre":
            data["nombre_proveedor"] = new_val.upper()
        elif field == "fecha":
            data["fecha_comprobante"] = new_val
        elif field == "metodo":
            data["metodo_pago"] = new_val.upper()
        elif field == "obs":
            data["observaciones"] = new_val
    except ValueError:
        await update.message.reply_text(
            f"⚠️ Valor inválido: `{new_val}`\nEscribe un número válido (sin símbolos).",
            parse_mode="Markdown",
        )
        return S_EDIT_VALUE

    if field in ("total", "itbis", "base", "propina"):
        total   = float(data.get("total_facturado") or 0)
        itbis   = float(data.get("itbis") or 0)
        propina = float(data.get("propina") or 0)
        base    = float(data.get("monto_sin_itbis") or 0)
        if field in ("total", "itbis"):
            base = round(total - itbis - propina, 2)
            data["monto_sin_itbis"] = base
        warns = []
        calc = round(base + itbis + propina, 2)
        if abs(calc - total) > 1.0:
            warns.append(f"⚠ Descuadre: {base:.2f}+{itbis:.2f}+{propina:.2f}={calc:.2f} ≠ {total:.2f}")
        existing = [w for w in (data.get("_warnings") or []) if "Descuadre" not in w]
        data["_warnings"] = existing + warns
        data["_needs_review"] = len(data["_warnings"]) > 0

    context.user_data["pending_invoice"] = data
    context.user_data.pop("edit_field", None)

    msg, kb = _review_msg_and_kb(context)
    await update.message.reply_text(msg, reply_markup=kb, parse_mode="Markdown")
    return S_CONFIRM


async def confirm_cancel(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Cancelar la factura actual."""
    query = update.callback_query
    await query.answer()
    _CLEANUP(context)
    await query.edit_message_text(
        "❌ Factura cancelada. Usa *Nueva factura* para empezar de nuevo.",
        parse_mode="Markdown",
    )
    return ConversationHandler.END


async def conv_cancel(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Cancelar toda la conversación (/cancelar)."""
    _CLEANUP(context)
    if update.message:
        await update.message.reply_text("Operación cancelada. Usa /nueva para empezar.")
    return ConversationHandler.END


# ──────────────────────────────────────────────────────────────
# COMANDOS DE CONSULTA (fuera del conversation handler)
# ──────────────────────────────────────────────────────────────

async def show_welcome(update: Update, context: ContextTypes.DEFAULT_TYPE,
                       newly_unlocked: bool = False):
    """Pantalla de bienvenida. Si newly_unlocked=True muestra el mensaje de acceso concedido."""
    mes = get_mes(context)
    if newly_unlocked:
        header = (
            "✅ *¡Acceso permitido!* Bienvenido a *Gabi* 🤖\n\n"
            "_Tu asistente virtual para el Formato 606 de la DGII_\n\n"
        )
    else:
        header = "🤖 *Gabi* — Asistente Formato 606 DGII\n\n"

    await update.message.reply_text(
        f"{header}"
        f"Mes activo: *{mes}*\n\n"
        f"👇 Usa los *botones de abajo* — no hace falta escribir nada:\n\n"
        f"🧾 *Nueva factura* — o envía directamente una foto\n"
        f"📊 *Resumen* — desglose de facturas por mes\n"
        f"📥 *Descargar Excel* — archivo 606 listo para entregar\n"
        f"⚠️ *Pendientes* — facturas con advertencias\n"
        f"❓ *Ayuda* — ver este menú nuevamente\n\n"
        f"_Avanzado:_ /lista  /borrar  /mes YYYY-MM",
        parse_mode="Markdown",
        reply_markup=main_menu_keyboard(),
    )


async def cmd_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_allowed(update): return
    await show_welcome(update, context)


async def cmd_ayuda(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_allowed(update): return
    await show_welcome(update, context)


async def auth_gate(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Puerta de entrada: exige la contraseña antes de usar el bot.
    Corre en el grupo -1 (antes que todo). Si el usuario no está desbloqueado:
      • si envía la contraseña → desbloquea y muestra el menú.
      • cualquier otra cosa    → pide la contraseña.
    Detiene la propagación con ApplicationHandlerStop en ambos casos."""
    user = update.effective_user
    if user is None:
        return  # updates sin usuario (no aplican)

    # La lista blanca de IDs (ALLOWED_USERS) sigue mandando si está configurada.
    if ALLOWED_USERS and user.id not in ALLOWED_USERS:
        raise ApplicationHandlerStop

    if context.user_data.get("unlocked"):
        return  # ya desbloqueado → dejar pasar a los demás handlers

    text = (update.message.text or "").strip() if (update.message and update.message.text) else ""
    if text == BOT_PASSWORD:
        context.user_data["unlocked"] = True
        await show_welcome(update, context, newly_unlocked=True)
        raise ApplicationHandlerStop

    # No desbloqueado y no es la contraseña → pedir clave y cortar.
    if update.message:
        await update.message.reply_text(
            "👋 ¡Hola! Soy *Gabi*, tu asistente virtual para el Formato 606 de la DGII 🤖\n\n"
            "🔒 Este bot es *privado*. Escribe la *contraseña de acceso* para continuar:",
            parse_mode="Markdown",
        )
    elif update.callback_query:
        await update.callback_query.answer(
            "🔒 Escribe la contraseña primero.", show_alert=True
        )
    raise ApplicationHandlerStop


async def _render_resumen(reply_to, mes: str, filtro: str = "ALL"):
    """Genera el resumen detallado del mes, factura por factura, con filtro de ubicación."""
    all_facturas = get_facturas(mes)
    facturas = [f for f in all_facturas if filtro == "ALL" or f.get("location") == filtro]

    filtro_label = f" — {filtro}" if filtro != "ALL" else ""
    if not facturas:
        await reply_to.reply_text(
            f"Sin facturas{filtro_label} para *{mes}*.", parse_mode="Markdown"
        )
        return

    rev  = sum(1 for f in facturas if f.get("needs_review"))
    qr_v = sum(1 for f in facturas if f.get("qr_verified"))
    header = (
        f"📊 *Resumen {mes}{filtro_label}*\n"
        f"_{len(facturas)} facturas  •  QR: {qr_v}  •  ⚠️ Revisión: {rev}_\n"
        f"{'─'*30}\n\n"
    )

    lines = []
    tot_t = tot_b = tot_i = tot_p = 0.0
    for i, f in enumerate(facturas, 1):
        nombre = (f.get("nombre") or "?")[:26]
        total  = float(f.get("total")   or 0)
        base   = float(f.get("base")    or 0)
        itbis  = float(f.get("itbis")   or 0)
        prop   = float(f.get("propina") or 0)
        if base == 0 and total > 0:
            base = round(total - itbis - prop, 2)
        tot_t += total; tot_b += base; tot_i += itbis; tot_p += prop

        fecha  = f.get("fecha_comp") or "—"
        ncf    = f.get("ncf") or "—"
        loc    = f.get("location") or "—"
        cat    = f.get("category") or "—"
        metodo = METODO_LABELS.get((f.get("metodo") or "").upper(), f.get("metodo") or "—")
        tg     = f.get("tipo_gasto") or ""
        tg_txt = (f"  🏷️ {tg} — {TIPO_GASTO_DICT.get(tg,'')}\n") if tg else ""
        adv    = "⚠️ " if f.get("needs_review") else ""
        qr     = "✅ " if f.get("qr_verified")  else ""

        line = (
            f"*{i}. {adv}{qr}{nombre}*\n"
            f"  📅 {fecha}  •  NCF: `{ncf}`\n"
            f"  📍 {loc}/{cat}  •  {metodo}\n"
            f"  Base: {base:,.2f}  ITBIS: {itbis:,.2f}"
            + (f"  Prop: {prop:,.2f}" if prop > 0 else "")
            + f"\n  💰 *RD$ {total:,.2f}*\n"
            + tg_txt
        )
        lines.append(line)

    totals_block = (
        f"{'═'*30}\n"
        f"*TOTALES — {len(facturas)} facturas*\n"
        f"Base sin ITBIS:  RD$ *{tot_b:,.2f}*\n"
        f"ITBIS:           RD$ *{tot_i:,.2f}*\n"
        + (f"Propina legal:   RD$ *{tot_p:,.2f}*\n" if tot_p > 0 else "")
        + f"{'━'*20}\n"
        f"*TOTAL: RD$ {tot_t:,.2f}*"
    )

    # Enviar en bloques si excede 4000 chars
    chunks, current = [], header
    for line in lines:
        if len(current) + len(line) > 3800:
            chunks.append(current)
            current = line
        else:
            current += line
    current += "\n" + totals_block
    chunks.append(current)

    for chunk in chunks:
        await reply_to.reply_text(chunk, parse_mode="Markdown")


async def cmd_resumen(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_allowed(update): return
    keyboard = InlineKeyboardMarkup([
        [InlineKeyboardButton("🌍 General (todas las ubicaciones)", callback_data="resfilter_ALL")],
        [
            InlineKeyboardButton("📍 Punta Cana",    callback_data="resfilter_Punta Cana"),
            InlineKeyboardButton("📍 Santo Domingo", callback_data="resfilter_Santo Domingo"),
        ],
    ])
    await update.message.reply_text(
        "📊 *Resumen — ¿qué ubicación?*", reply_markup=keyboard, parse_mode="Markdown"
    )


async def on_pick_resfilter(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Seleccionada la ubicación → mostrar selector de mes."""
    query = update.callback_query
    await query.answer()
    filtro = query.data.replace("resfilter_", "")
    context.user_data["res_filtro"] = filtro
    kb = mes_picker("res")
    if kb is None:
        await query.edit_message_text("Aún no hay facturas guardadas.")
        return
    etiq = "todas las ubicaciones" if filtro == "ALL" else filtro
    await query.edit_message_text(
        f"📊 *{etiq}* — ¿De qué mes quieres el resumen?",
        reply_markup=kb, parse_mode="Markdown"
    )


async def on_pick_resumen(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    mes = query.data.replace("res_", "")
    filtro = context.user_data.pop("res_filtro", "ALL")
    context.user_data["mes_activo"] = mes
    etiq = "" if filtro == "ALL" else f" — {filtro}"
    await query.edit_message_text(f"📊 Generando resumen *{mes}{etiq}*…", parse_mode="Markdown")
    await _render_resumen(query.message, mes, filtro)


async def cmd_lista(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_allowed(update): return
    mes = get_mes(context)
    facturas = get_facturas(mes)
    if not facturas:
        await update.message.reply_text(f"Sin facturas para *{mes}*.", parse_mode="Markdown")
        return

    lines = [f"📋 *{mes}* — {len(facturas)} facturas\n"]
    for i, f in enumerate(facturas, 1):
        qr  = "✓" if f["qr_verified"] else " "
        rev = "⚠" if f["needs_review"] else " "
        nombre = (f["nombre"] or f["filename"] or "?")[:22]
        loc = (f.get("location") or "")[:2]
        cat = (f.get("category") or "")[:1]
        lines.append(f"{i:>2}. {qr}{rev} {nombre:<22} RD${f['total']:>9,.2f}  {loc}/{cat}")

    text = "\n".join(lines)
    if len(text) > 3800:
        text = text[:3800] + "\n…(usa /exportar para el Excel completo)"
    await update.message.reply_text(f"```\n{text}\n```", parse_mode="Markdown")


async def cmd_pendientes(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_allowed(update): return
    mes = get_mes(context)
    facturas = [f for f in get_facturas(mes) if f["needs_review"]]
    if not facturas:
        await update.message.reply_text("✅ Sin advertencias pendientes.")
        return

    lines = [f"⚠️ *Advertencias {mes}*\n"]
    for f in facturas:
        adv = json.loads(f["advertencias"] or "[]")
        lines.append(f"• *{f['nombre'] or f['ncf']}* — RD${f['total']:,.2f}")
        for a in adv:
            lines.append(f"  {a}")
    await update.message.reply_text("\n".join(lines), parse_mode="Markdown")


async def _render_exportar(reply_to, mes: str):
    """Genera el Excel del mes indicado y lo envía como documento."""
    facturas = get_facturas(mes)
    if not facturas:
        await reply_to.reply_text(f"Sin facturas para *{mes}*.", parse_mode="Markdown")
        return

    await reply_to.reply_text(f"⏳ Generando Excel *{mes}*...", parse_mode="Markdown")
    xlsx = build_excel(facturas, mes)
    await reply_to.reply_document(
        document=io.BytesIO(xlsx),
        filename=f"606_{mes}.xlsx",
        caption=f"✅ Formato 606 — {mes} — {len(facturas)} facturas",
    )


async def cmd_exportar(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_allowed(update): return
    kb = mes_picker("exp")
    if kb is None:
        await update.message.reply_text("Aún no hay facturas guardadas.")
        return
    await update.message.reply_text(
        "📥 ¿Qué mes quieres exportar?", reply_markup=kb
    )


async def on_pick_exportar(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    mes = query.data.replace("exp_", "")
    context.user_data["mes_activo"] = mes
    await query.edit_message_text(f"📥 Exportando *{mes}*…", parse_mode="Markdown")
    await _render_exportar(query.message, mes)


async def cmd_borrar(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_allowed(update): return
    mes = get_mes(context)
    deleted = delete_last(mes)
    if deleted:
        await update.message.reply_text(
            f"🗑 Eliminada: *{deleted['nombre'] or deleted['ncf']}*  RD${deleted['total']:,.2f}",
            parse_mode="Markdown",
        )
    else:
        await update.message.reply_text("No hay facturas para eliminar.")


async def cmd_mes(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_allowed(update): return
    args = context.args
    if not args or not re.match(r"^\d{4}-\d{2}$", args[0]):
        await update.message.reply_text("Formato: /mes YYYY-MM  (ej: /mes 2026-06)")
        return
    context.user_data["mes_activo"] = args[0]
    n = len(get_facturas(args[0]))
    await update.message.reply_text(
        f"📅 Mes activo: *{args[0]}* ({n} facturas)", parse_mode="Markdown"
    )


# ──────────────────────────────────────────────────────────────
# EXCEL 606
# ──────────────────────────────────────────────────────────────

def build_excel(facturas: list[dict], mes: str) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    def fill(c): return PatternFill(start_color=c, end_color=c, fill_type="solid")
    def font(bold=False, color="000000", size=9):
        return Font(name="Calibri", bold=bold, color=color, size=size)
    def bdr():
        s = Side(style="thin", color="AAAAAA")
        return Border(left=s, right=s, top=s, bottom=s)
    def aln(h="left", wrap=False):
        return Alignment(horizontal=h, vertical="center", wrap_text=wrap)

    wb = Workbook()
    ws = wb.active
    ws.title = "Formato 606"

    ws.merge_cells("A1:V1")
    c = ws["A1"]
    c.value = f"FORMATO 606 — COMPRAS — {mes}"
    c.font = Font(name="Calibri", bold=True, color="FFFFFF", size=13)
    c.fill = fill("1F4E79")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:V2")
    c = ws["A2"]
    rev = sum(1 for f in facturas if f.get("needs_review"))
    c.value = (f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}  |  "
               f"Facturas: {len(facturas)}  |  Revisión pendiente: {rev}")
    c.font = Font(name="Calibri", italic=True, color="FFFFFF", size=9)
    c.fill = fill("2E75B6")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 16

    COLS = [
        ("No.",6),("RNC/Cédula",16),("T.ID",5),("Nombre Proveedor",28),
        ("NCF",17),("T.CF",6),("Fec.Comp.",12),("Día\nPago",7),
        ("Base\nSin ITBIS",13),("ITBIS\nFact.",12),("Total\nFact.",13),
        ("Propina\n10%",11),("Mét.\nPago",7),("ITBIS\nRet.",10),
        ("ITBIS\nPerc.",10),("T.Ret\nISR",7),("Mto.\nRenta",11),
        ("Mto.\nServ.",12),("Mto.\nBienes",11),
        ("Tipo\nGasto",9),
        ("Observaciones",35),
        ("Subido\npor",16),
    ]
    for col, (hdr, w) in enumerate(COLS, 1):
        c = ws.cell(row=3, column=col, value=hdr)
        c.font = font(bold=True, color="FFFFFF")
        c.fill = fill("2E75B6")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = bdr()
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[3].height = 32

    totals = dict(base=0.0, itbis=0.0, total=0.0, propina=0.0)

    def tipo_cf(ncf):
        n = (ncf or "").upper()
        if n.startswith(("B01","E31")): return "01"
        if n.startswith(("B02","E32")): return "02"
        if n.startswith(("B14","E34")): return "14"
        if n.startswith(("B11","E41")): return "11"
        return "02"

    def met_code(m):
        m = (m or "").upper()
        if any(x in m for x in ("EFECTIVO","CASH","METALICO","METÁLICO")): return "01"
        if "CHEQUE" in m: return "02"
        if any(x in m for x in ("TRANSFER","DEPOSITO")): return "03"
        if any(x in m for x in ("CREDITO","CRÉDITO","CREDIT")): return "04"
        if any(x in m for x in ("DEBITO","DÉBITO","DEBIT","PIN")): return "05"
        return "01"

    def rnc_tipo(rnc):
        return "2" if len(re.sub(r"\D","",str(rnc or "")))==11 else "1"

    # Orden cronológico por fecha de comprobante; a igual fecha, por orden de
    # envío (id ascendente). No se agrupa por proveedor.
    facturas = sorted(
        facturas,
        key=lambda f: (f.get("fecha_comp") or "", f.get("id") or 0),
    )

    for seq, fac in enumerate(facturas, 1):
        row = seq + 3
        is_qr   = bool(fac.get("qr_verified"))
        needs   = bool(fac.get("needs_review"))

        if is_qr:   rf = fill("E2EFDA")
        elif needs: rf = fill("FFF2CC")
        elif seq%2==0: rf = fill("DEEAF1")
        else: rf = PatternFill()

        rnc   = re.sub(r"\D","",str(fac.get("rnc") or ""))
        total = fac.get("total",0) or 0
        itbis = fac.get("itbis",0) or 0
        base  = fac.get("base",0) or 0
        prop  = fac.get("propina",0) or 0
        if base==0 and total>0:
            base = round(total-itbis-prop, 2)

        totals["base"]    += base
        totals["itbis"]   += itbis
        totals["total"]   += total
        totals["propina"] += prop

        fp = fac.get("fecha_pago") or fac.get("fecha_comp") or ""
        if len(str(fp))==10 and "-" in str(fp):
            fp = str(fp).split("-")[2].lstrip("0") or "0"

        obs = fac.get("observaciones") or ""
        adv = json.loads(fac.get("advertencias") or "[]")
        if adv: obs = (obs+" | "+" | ".join(adv)).strip(" |")
        if is_qr: obs = ("✓ QR. "+obs).strip()
        loc = fac.get("location",""); cat = fac.get("category","")
        if loc or cat:
            tag = f"[{loc}/{cat}]"
            obs = (tag+" "+obs).strip() if obs else tag

        tg = fac.get("tipo_gasto") or "02"
        tg_label = TIPO_GASTO_DICT.get(tg, tg)
        vals = [
            seq, rnc, rnc_tipo(rnc),
            fac.get("nombre",""),
            fac.get("ncf",""),
            tipo_cf(fac.get("ncf","")),
            fac.get("fecha_comp",""),
            fp,
            f"{base:.2f}", f"{itbis:.2f}", f"{total:.2f}", f"{prop:.2f}",
            met_code(fac.get("metodo","")),
            "0.00","0.00","","0.00",
            f"{base:.2f}","0.00",
            f"{tg} – {tg_label}",
            obs,
            fac.get("usuario","") or "",
        ]
        CENTER={2,3,5,6,7,8,13,16}; RIGHT={1,9,10,11,12,14,15,17,18,19}
        for col,val in enumerate(vals,1):
            c = ws.cell(row=row, column=col, value=val)
            c.font = font(size=9); c.fill=rf; c.border=bdr()
            if col in CENTER: c.alignment=aln("center")
            elif col in RIGHT: c.alignment=aln("right")
            else: c.alignment=aln("left")
        ws.row_dimensions[row].height=16

    tr = len(facturas)+4
    ws.merge_cells(f"A{tr}:H{tr}")
    c=ws[f"A{tr}"]; c.value="TOTALES"
    c.font=font(bold=True,size=10); c.fill=fill("FFC000")
    c.alignment=Alignment(horizontal="center",vertical="center")
    for off,key in enumerate(["base","itbis","total","propina"]):
        c=ws.cell(row=tr,column=9+off,value=f"{totals[key]:.2f}")
        c.font=font(bold=True,size=10); c.fill=fill("FFC000"); c.alignment=aln("right")
    for col in range(13,23):  # A-V = 22 columns
        ws.cell(row=tr,column=col).fill=fill("FFC000")
    ws.row_dimensions[tr].height=20
    ws.freeze_panes="A4"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ──────────────────────────────────────────────────────────────
# MAIN
# ──────────────────────────────────────────────────────────────

def main():
    init_db()
    app = Application.builder().token(BOT_TOKEN).build()

    # Puerta de contraseña: corre primero (grupo -1) en cada update.
    app.add_handler(TypeHandler(Update, auth_gate), group=-1)

    # Conversation handler for the guided capture flow
    conv = ConversationHandler(
        entry_points=[
            CommandHandler("nueva", start_flow),
            MessageHandler(filters.Regex(f"^{re.escape(BTN_NUEVA)}$"), start_flow),
            MessageHandler(filters.PHOTO & filters.ChatType.PRIVATE, start_flow),
        ],
        states={
            S_PHOTO: [
                MessageHandler(filters.PHOTO, got_photo),
                CallbackQueryHandler(otra_factura, pattern="^otra_factura$"),
                CallbackQueryHandler(fin_lote,     pattern="^fin_lote$"),
            ],
            S_LOCATION: [
                CallbackQueryHandler(got_location, pattern="^loc_"),
            ],
            S_CATEGORY: [
                CallbackQueryHandler(got_category,     pattern="^cat_"),
                CallbackQueryHandler(back_to_location, pattern="^back_to_location$"),
            ],
            S_CONFIRM: [
                CallbackQueryHandler(confirm_accept,        pattern="^confirm_accept$"),
                CallbackQueryHandler(confirm_edit,          pattern="^confirm_edit$"),
                CallbackQueryHandler(confirm_cancel,        pattern="^confirm_cancel$"),
                CallbackQueryHandler(on_confirm_change_mes, pattern="^confirm_change_mes$"),
                CallbackQueryHandler(on_cmes_picked,        pattern="^cmes_"),
            ],
            S_EDIT_SELECT: [
                CallbackQueryHandler(edit_field_selected,  pattern="^edit_"),
                CallbackQueryHandler(edit_tipo_gasto_pick, pattern="^edittg_"),
            ],
            S_EDIT_VALUE: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, edit_value_received),
            ],
        },
        fallbacks=[
            CommandHandler("cancelar", conv_cancel),
            MessageHandler(filters.Regex(f"^{re.escape(BTN_NUEVA)}$"), start_flow),
        ],
        allow_reentry=True,
        per_user=True,
        per_chat=True,
    )

    app.add_handler(conv)

    # Standalone commands (outside conversation)
    app.add_handler(CommandHandler("start",      cmd_start))
    app.add_handler(CommandHandler("resumen",    cmd_resumen))
    app.add_handler(CommandHandler("lista",      cmd_lista))
    app.add_handler(CommandHandler("pendientes", cmd_pendientes))
    app.add_handler(CommandHandler("exportar",   cmd_exportar))
    app.add_handler(CommandHandler("borrar",     cmd_borrar))
    app.add_handler(CommandHandler("mes",        cmd_mes))
    app.add_handler(CommandHandler("ayuda",      cmd_ayuda))

    # Botones del menú fijo → mismos comandos (sin escribir nada)
    app.add_handler(MessageHandler(filters.Regex(f"^{re.escape(BTN_RESUMEN)}$"),    cmd_resumen))
    app.add_handler(MessageHandler(filters.Regex(f"^{re.escape(BTN_EXPORTAR)}$"),   cmd_exportar))
    app.add_handler(MessageHandler(filters.Regex(f"^{re.escape(BTN_PENDIENTES)}$"), cmd_pendientes))
    app.add_handler(MessageHandler(filters.Regex(f"^{re.escape(BTN_AYUDA)}$"),      cmd_ayuda))

    # Selección de ubicación y mes para /resumen y /exportar
    app.add_handler(CallbackQueryHandler(on_pick_resfilter, pattern="^resfilter_"))
    app.add_handler(CallbackQueryHandler(on_pick_resumen,   pattern="^res_"))
    app.add_handler(CallbackQueryHandler(on_pick_exportar,  pattern="^exp_"))

    log.info("Bot iniciado con flujo guiado.")
    app.run_polling(drop_pending_updates=True)


if __name__ == "__main__":
    main()
