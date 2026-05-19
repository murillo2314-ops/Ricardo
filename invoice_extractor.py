#!/usr/bin/env python3
"""
Extractor de Facturas para Formato 606 DGII - República Dominicana
Usa Claude Vision API para extraer datos de facturas en imágenes/PDFs dentro de un ZIP.
"""

import argparse
import base64
import io
import json
import os
import sys
import zipfile
from datetime import datetime
from pathlib import Path

import anthropic
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side
)
from openpyxl.utils import get_column_letter

load_dotenv()

EXTRACTION_PROMPT = """Eres un experto en facturación dominicana y fiscalización de la DGII (Dirección General de Impuestos Internos) de República Dominicana.

Tu tarea es analizar imágenes de facturas y comprobantes fiscales dominicanos y extraer los datos exactos que aparecen en ellos.

CAMPOS A EXTRAER:
- rnc: RNC o Cédula del proveedor/emisor (solo dígitos, sin guiones ni espacios). Puede ser 9 dígitos (RNC empresa) u 11 dígitos (cédula persona).
- ncf: Número de Comprobante Fiscal (NCF). Formato B seguido de 2 dígitos y 8 dígitos más (ej: B0100000001) o E seguido de 2 dígitos y 10 dígitos (ej: E310000000001). Puede aparecer como "NCF", "No. Comprobante", "Comprobante Fiscal", etc.
- fecha_comprobante: Fecha de emisión del comprobante en formato YYYY-MM-DD. Si no aparece explícitamente, usa la fecha de la factura.
- fecha_pago: Fecha de pago o vencimiento en formato YYYY-MM-DD. Si no está, usa la misma que fecha_comprobante.
- total_facturado: Monto total de la factura como número decimal (incluyendo ITBIS y propina si aplica).
- itbis: Monto del ITBIS (18%) como número decimal. Si aparece como "ITBIS", "IVA 18%", "Tax", etc.
- monto_sin_itbis: Monto antes de ITBIS como número decimal. Si no está explícito, calcular como total_facturado - itbis - propina.
- metodo_pago: Método de pago. Valores posibles: "EFECTIVO", "TARJETA_CREDITO", "TARJETA_DEBITO", "TRANSFERENCIA", "CHEQUE", "CREDITO". Si no está claro, pon "EFECTIVO".
- propina: Monto de propina legal (10%) como número decimal. Si no hay propina, pon 0.
- nombre_proveedor: Nombre o razón social del emisor/proveedor de la factura.
- tipo_comprobante: Tipo de NCF basado en los primeros caracteres del NCF:
  - B01 o E31 → "CREDITO_FISCAL"
  - B02 o E32 → "CONSUMIDOR_FINAL"
  - B14 o E34 → "REGIMEN_ESPECIAL"
  - B15 o E35 → "GUBERNAMENTAL"
  - B11 o E41 → "GASTOS_MENORES"
  - Otro → "CONSUMIDOR_FINAL"
- observaciones: Cualquier dato relevante adicional o nota sobre la factura.

REGLAS IMPORTANTES:
1. Si un campo no aparece claramente en la imagen, usa null para ese campo.
2. Para montos, usa solo números decimales sin símbolos de moneda ni comas como separadores de miles (ej: 1500.00 no $1,500.00).
3. Para RNC, elimina todos los guiones y espacios.
4. Si el total facturado no está claro, súmalo: monto_sin_itbis + itbis + propina.
5. Para el tipo_comprobante, analiza el NCF completo para determinarlo.
6. Responde ÚNICAMENTE con un objeto JSON válido, sin texto adicional, sin explicaciones, sin markdown.

FORMATO DE RESPUESTA (solo JSON):
{
  "rnc": "123456789",
  "ncf": "B0100000001",
  "fecha_comprobante": "2024-01-15",
  "fecha_pago": "2024-01-15",
  "total_facturado": 1180.00,
  "itbis": 180.00,
  "monto_sin_itbis": 1000.00,
  "metodo_pago": "TARJETA_CREDITO",
  "propina": 0,
  "nombre_proveedor": "Empresa Ejemplo SRL",
  "tipo_comprobante": "CREDITO_FISCAL",
  "observaciones": null
}"""

SUPPORTED_IMAGE_TYPES = {
    '.jpg': 'image/jpeg',
    '.jpeg': 'image/jpeg',
    '.png': 'image/png',
    '.gif': 'image/gif',
    '.webp': 'image/webp',
}


def encode_image_bytes_to_base64(image_bytes: bytes) -> str:
    return base64.standard_b64encode(image_bytes).decode('utf-8')


def pdf_to_images(pdf_bytes: bytes) -> list[tuple[bytes, str]]:
    """Convert PDF pages to PNG images. Returns list of (image_bytes, 'image/png')."""
    try:
        import fitz  # PyMuPDF
    except ImportError:
        print("  [AVISO] PyMuPDF no instalado. No se pueden procesar PDFs.")
        return []

    images = []
    doc = fitz.open(stream=pdf_bytes, filetype="pdf")
    for page_num in range(len(doc)):
        page = doc.load_page(page_num)
        mat = fitz.Matrix(2.0, 2.0)  # 2x resolution for better quality
        pix = page.get_pixmap(matrix=mat)
        img_bytes = pix.tobytes("png")
        images.append((img_bytes, 'image/png'))
    doc.close()
    return images


def extract_invoice_data(client: anthropic.Anthropic, image_data: str, media_type: str, filename: str) -> dict:
    """Call Claude API to extract invoice data from a base64-encoded image."""
    try:
        with client.messages.stream(
            model="claude-opus-4-7",
            max_tokens=2048,
            thinking={"type": "adaptive"},
            system=[{
                "type": "text",
                "text": EXTRACTION_PROMPT,
                "cache_control": {"type": "ephemeral"}
            }],
            messages=[{
                "role": "user",
                "content": [
                    {
                        "type": "image",
                        "source": {
                            "type": "base64",
                            "media_type": media_type,
                            "data": image_data,
                        }
                    },
                    {
                        "type": "text",
                        "text": f"Extrae los datos de esta factura (archivo: {filename}). Responde solo con JSON válido."
                    }
                ]
            }]
        ) as stream:
            response = stream.get_final_message()

        # Extract text content from response
        text_content = ""
        for block in response.content:
            if block.type == "text":
                text_content = block.text
                break

        # Clean markdown code blocks if present
        text_content = text_content.strip()
        if text_content.startswith("```"):
            lines = text_content.split('\n')
            # Remove first and last lines (```json and ```)
            inner = '\n'.join(lines[1:])
            if inner.endswith("```"):
                inner = inner[:-3]
            text_content = inner.strip()

        data = json.loads(text_content)
        data['_filename'] = filename
        data['_error'] = None
        return data

    except json.JSONDecodeError as e:
        print(f"  [ERROR] No se pudo parsear JSON de {filename}: {e}")
        return {'_filename': filename, '_error': f'JSON parse error: {e}'}
    except Exception as e:
        print(f"  [ERROR] Fallo al procesar {filename}: {e}")
        return {'_filename': filename, '_error': str(e)}


def clean_rnc(rnc) -> str:
    """Remove all non-digit characters from RNC."""
    if rnc is None:
        return ""
    return ''.join(c for c in str(rnc) if c.isdigit())


def determine_tipo_id(rnc: str) -> str:
    """Return '1' for 9-digit RNC (empresa), '2' for 11-digit cédula."""
    digits = clean_rnc(rnc)
    if len(digits) == 11:
        return "2"
    return "1"


def format_amount(value) -> str:
    """Format numeric value to 2 decimal places."""
    if value is None:
        return "0.00"
    try:
        return f"{float(value):.2f}"
    except (ValueError, TypeError):
        return "0.00"


def get_tipo_comprobante_code(tipo: str) -> str:
    """Map tipo_comprobante string to DGII numeric code."""
    mapping = {
        "CREDITO_FISCAL": "01",
        "CONSUMIDOR_FINAL": "02",
        "REGIMEN_ESPECIAL": "14",
        "GUBERNAMENTAL": "15",
        "GASTOS_MENORES": "11",
    }
    return mapping.get(tipo, "02")


def get_metodo_pago_code(metodo: str) -> str:
    """Map metodo_pago string to DGII numeric code."""
    if metodo is None:
        return "01"
    metodo = str(metodo).upper()
    if "EFECTIVO" in metodo:
        return "01"
    elif "CHEQUE" in metodo:
        return "02"
    elif "TRANSFERENCIA" in metodo or "DEPOSITO" in metodo:
        return "03"
    elif "TARJETA_CREDITO" in metodo or "CREDITO" in metodo:
        return "04"
    elif "TARJETA_DEBITO" in metodo or "DEBITO" in metodo:
        return "05"
    elif "CREDITO" in metodo:
        return "06"
    return "01"


def create_excel_report(invoices: list[dict], output_path: str):
    """Create an Excel workbook with DGII 606 format and raw data sheets."""
    wb = Workbook()

    # --- Sheet 1: Formato 606 ---
    ws606 = wb.active
    ws606.title = "Formato 606"

    # Color definitions
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    subheader_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    alt_fill = PatternFill(start_color="DEEAF1", end_color="DEEAF1", fill_type="solid")
    error_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    total_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")

    white_font = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    normal_font = Font(name="Calibri", size=10)
    bold_font = Font(name="Calibri", bold=True, size=10)
    error_font = Font(name="Calibri", size=10, color="CC0000")

    thin_side = Side(style="thin", color="AAAAAA")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    thick_side = Side(style="medium", color="2E75B6")
    thick_border = Border(left=thick_side, right=thick_side, top=thick_side, bottom=thick_side)

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center")
    right = Alignment(horizontal="right", vertical="center")

    # Title row
    ws606.merge_cells("A1:T1")
    title_cell = ws606["A1"]
    title_cell.value = "FORMATO 606 - COMPRAS DE BIENES Y SERVICIOS"
    title_cell.font = header_font
    title_cell.fill = header_fill
    title_cell.alignment = center
    ws606.row_dimensions[1].height = 30

    # Subtitle row
    ws606.merge_cells("A2:T2")
    subtitle_cell = ws606["A2"]
    subtitle_cell.value = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}  |  Total Facturas: {len([i for i in invoices if not i.get('_error')])}"
    subtitle_cell.font = Font(name="Calibri", italic=True, color="FFFFFF", size=9)
    subtitle_cell.fill = subheader_fill
    subtitle_cell.alignment = center
    ws606.row_dimensions[2].height = 20

    # Column headers (row 3)
    columns_606 = [
        ("No.", 5),
        ("RNC/Cédula\nProveedor", 18),
        ("Tipo\nID", 6),
        ("Nombre Proveedor", 30),
        ("NCF", 18),
        ("Tipo\nComprobante", 8),
        ("Fecha\nComprobante", 13),
        ("Fecha\nPago", 13),
        ("Monto\nSin ITBIS", 14),
        ("ITBIS\nFacturado", 14),
        ("Total\nFacturado", 14),
        ("Propina\nLegal 10%", 12),
        ("Método\nde Pago", 8),
        ("ITBIS\nRetenido", 12),
        ("ITBIS\nPercibido", 12),
        ("Tipo\nRet. ISR", 8),
        ("Monto\nRet. Renta", 12),
        ("Monto\nServicios", 14),
        ("Monto\nBienes", 14),
        ("Observaciones", 25),
    ]

    for col_idx, (header, width) in enumerate(columns_606, start=1):
        cell = ws606.cell(row=3, column=col_idx, value=header)
        cell.font = white_font
        cell.fill = subheader_fill
        cell.alignment = center
        cell.border = thin_border
        ws606.column_dimensions[get_column_letter(col_idx)].width = width
    ws606.row_dimensions[3].height = 35

    # Data rows
    totals = {
        'monto_sin_itbis': 0.0,
        'itbis': 0.0,
        'total_facturado': 0.0,
        'propina': 0.0,
    }

    valid_row = 0
    for row_idx, invoice in enumerate(invoices, start=1):
        excel_row = row_idx + 3
        is_error = bool(invoice.get('_error'))
        fill = error_fill if is_error else (alt_fill if row_idx % 2 == 0 else PatternFill())
        font = error_font if is_error else normal_font

        if is_error:
            row_data = [
                row_idx,
                "",
                "",
                f"ERROR: {invoice['_filename']}",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                invoice.get('_error', 'Error desconocido'),
            ]
        else:
            valid_row += 1
            rnc = clean_rnc(invoice.get('rnc'))
            total = float(invoice.get('total_facturado') or 0)
            itbis = float(invoice.get('itbis') or 0)
            propina = float(invoice.get('propina') or 0)
            monto_sin_itbis = float(invoice.get('monto_sin_itbis') or 0)

            # Recalculate if needed
            if total == 0 and (monto_sin_itbis > 0 or itbis > 0):
                total = monto_sin_itbis + itbis + propina
            if monto_sin_itbis == 0 and total > 0:
                monto_sin_itbis = total - itbis - propina

            totals['monto_sin_itbis'] += monto_sin_itbis
            totals['itbis'] += itbis
            totals['total_facturado'] += total
            totals['propina'] += propina

            row_data = [
                valid_row,
                rnc,
                determine_tipo_id(rnc),
                invoice.get('nombre_proveedor') or "",
                invoice.get('ncf') or "",
                get_tipo_comprobante_code(invoice.get('tipo_comprobante') or ""),
                invoice.get('fecha_comprobante') or "",
                invoice.get('fecha_pago') or "",
                format_amount(monto_sin_itbis),
                format_amount(itbis),
                format_amount(total),
                format_amount(propina),
                get_metodo_pago_code(invoice.get('metodo_pago')),
                "0.00",  # ITBIS Retenido
                "0.00",  # ITBIS Percibido
                "",       # Tipo Ret. ISR
                "0.00",  # Monto Ret. Renta
                format_amount(monto_sin_itbis),  # Monto Servicios (default to full)
                "0.00",  # Monto Bienes
                invoice.get('observaciones') or "",
            ]

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws606.cell(row=excel_row, column=col_idx, value=value)
            cell.font = font
            cell.fill = fill
            cell.border = thin_border
            # Right-align numeric columns
            if col_idx in (1, 9, 10, 11, 12, 14, 15, 17, 18, 19):
                cell.alignment = right
            elif col_idx in (2, 3, 5, 6, 7, 8, 13, 16):
                cell.alignment = center
            else:
                cell.alignment = left

        ws606.row_dimensions[excel_row].height = 18

    # Totals row
    totals_row = len(invoices) + 4
    ws606.merge_cells(f"A{totals_row}:H{totals_row}")
    total_label = ws606[f"A{totals_row}"]
    total_label.value = "TOTALES"
    total_label.font = bold_font
    total_label.fill = total_fill
    total_label.alignment = center
    total_label.border = thick_border

    totals_values = [
        totals['monto_sin_itbis'],
        totals['itbis'],
        totals['total_facturado'],
        totals['propina'],
    ]
    for col_offset, val in enumerate(totals_values):
        col_idx = 9 + col_offset
        cell = ws606.cell(row=totals_row, column=col_idx, value=format_amount(val))
        cell.font = bold_font
        cell.fill = total_fill
        cell.alignment = right
        cell.border = thick_border

    # Empty remaining total cells
    for col_idx in range(13, 21):
        cell = ws606.cell(row=totals_row, column=col_idx, value="")
        cell.fill = total_fill
        cell.border = thick_border

    ws606.row_dimensions[totals_row].height = 22
    ws606.freeze_panes = "A4"

    # --- Sheet 2: Datos Extraídos (raw data for review) ---
    ws_raw = wb.create_sheet("Datos Extraídos")
    raw_headers = [
        "Archivo", "RNC/Cédula", "NCF", "Nombre Proveedor",
        "Fecha Comprobante", "Fecha Pago", "Monto Sin ITBIS",
        "ITBIS", "Propina", "Total Facturado",
        "Método de Pago", "Tipo Comprobante", "Observaciones", "Error"
    ]
    raw_widths = [30, 15, 18, 35, 16, 16, 15, 15, 12, 15, 18, 18, 30, 40]

    for col_idx, (header, width) in enumerate(zip(raw_headers, raw_widths), start=1):
        cell = ws_raw.cell(row=1, column=col_idx, value=header)
        cell.font = white_font
        cell.fill = subheader_fill
        cell.alignment = center
        cell.border = thin_border
        ws_raw.column_dimensions[get_column_letter(col_idx)].width = width
    ws_raw.row_dimensions[1].height = 25

    for row_idx, invoice in enumerate(invoices, start=2):
        is_error = bool(invoice.get('_error'))
        fill = error_fill if is_error else (alt_fill if row_idx % 2 == 0 else PatternFill())
        raw_row = [
            invoice.get('_filename', ''),
            clean_rnc(invoice.get('rnc')) if not is_error else '',
            invoice.get('ncf', '') if not is_error else '',
            invoice.get('nombre_proveedor', '') if not is_error else '',
            invoice.get('fecha_comprobante', '') if not is_error else '',
            invoice.get('fecha_pago', '') if not is_error else '',
            format_amount(invoice.get('monto_sin_itbis')) if not is_error else '',
            format_amount(invoice.get('itbis')) if not is_error else '',
            format_amount(invoice.get('propina')) if not is_error else '',
            format_amount(invoice.get('total_facturado')) if not is_error else '',
            invoice.get('metodo_pago', '') if not is_error else '',
            invoice.get('tipo_comprobante', '') if not is_error else '',
            invoice.get('observaciones', '') if not is_error else '',
            invoice.get('_error', '') or '',
        ]
        for col_idx, value in enumerate(raw_row, start=1):
            cell = ws_raw.cell(row=row_idx, column=col_idx, value=value)
            cell.font = error_font if is_error else normal_font
            cell.fill = fill
            cell.border = thin_border
            cell.alignment = left
        ws_raw.row_dimensions[row_idx].height = 16

    ws_raw.freeze_panes = "A2"

    wb.save(output_path)
    print(f"\n[OK] Reporte Excel guardado en: {output_path}")


def process_zip_file(zip_path: str, client: anthropic.Anthropic) -> list[dict]:
    """Open ZIP, process each supported image/PDF, return list of extracted invoice data."""
    invoices = []

    with zipfile.ZipFile(zip_path, 'r') as zf:
        all_names = zf.namelist()
        # Filter only supported files, skip hidden/system files
        supported_files = [
            name for name in all_names
            if not os.path.basename(name).startswith(('.', '__'))
            and (
                Path(name).suffix.lower() in SUPPORTED_IMAGE_TYPES
                or Path(name).suffix.lower() == '.pdf'
            )
        ]

        if not supported_files:
            print("[ERROR] No se encontraron imágenes ni PDFs en el ZIP.")
            return []

        print(f"[INFO] Encontrados {len(supported_files)} archivo(s) para procesar.")

        for file_name in supported_files:
            filename = os.path.basename(file_name)
            ext = Path(filename).suffix.lower()
            print(f"\n-> Procesando: {filename}")

            file_bytes = zf.read(file_name)

            if ext == '.pdf':
                print("   Convirtiendo PDF a imágenes...")
                pages = pdf_to_images(file_bytes)
                if not pages:
                    invoices.append({'_filename': filename, '_error': 'No se pudo convertir PDF a imagen'})
                    continue

                # Process each page, prefer the one with most data
                page_results = []
                for page_num, (img_bytes, media_type) in enumerate(pages, start=1):
                    page_filename = f"{filename} (pág. {page_num})"
                    print(f"   Extrayendo datos de página {page_num}/{len(pages)}...")
                    img_b64 = encode_image_bytes_to_base64(img_bytes)
                    result = extract_invoice_data(client, img_b64, media_type, page_filename)
                    page_results.append(result)

                # Pick best result: prefer one with NCF
                best = next((r for r in page_results if r.get('ncf') and not r.get('_error')), None)
                if best is None:
                    best = page_results[0] if page_results else {'_filename': filename, '_error': 'Sin resultados'}
                best['_filename'] = filename
                invoices.append(best)

            else:
                media_type = SUPPORTED_IMAGE_TYPES[ext]
                img_b64 = encode_image_bytes_to_base64(file_bytes)
                print(f"   Extrayendo datos de factura...")
                result = extract_invoice_data(client, img_b64, media_type, filename)
                invoices.append(result)

    return invoices


def main():
    parser = argparse.ArgumentParser(
        description="Extrae datos de facturas dominicanas y genera Formato 606 DGII en Excel.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Ejemplos:
  python invoice_extractor.py facturas.zip
  python invoice_extractor.py facturas.zip -o reporte_606.xlsx

Requiere la variable de entorno ANTHROPIC_API_KEY o un archivo .env con ella.
        """
    )
    parser.add_argument("zip_path", help="Ruta al archivo ZIP con las facturas (imágenes y/o PDFs)")
    parser.add_argument(
        "-o", "--output",
        default=None,
        help="Ruta del archivo Excel de salida (por defecto: 606_<fecha>.xlsx)"
    )
    args = parser.parse_args()

    api_key = os.environ.get("ANTHROPIC_API_KEY")
    if not api_key:
        print("[ERROR] No se encontró ANTHROPIC_API_KEY. Define esta variable de entorno o crea un archivo .env")
        sys.exit(1)

    if not os.path.isfile(args.zip_path):
        print(f"[ERROR] Archivo ZIP no encontrado: {args.zip_path}")
        sys.exit(1)

    output_path = args.output
    if not output_path:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = f"606_{timestamp}.xlsx"

    print("=" * 60)
    print("  EXTRACTOR DE FACTURAS - FORMATO 606 DGII")
    print("=" * 60)
    print(f"ZIP: {args.zip_path}")
    print(f"Salida: {output_path}")
    print()

    client = anthropic.Anthropic(api_key=api_key)

    invoices = process_zip_file(args.zip_path, client)

    if not invoices:
        print("[ERROR] No se procesaron facturas.")
        sys.exit(1)

    success = sum(1 for i in invoices if not i.get('_error'))
    errors = len(invoices) - success
    print(f"\n[RESUMEN] Procesadas: {len(invoices)} | Exitosas: {success} | Con error: {errors}")

    create_excel_report(invoices, output_path)

    print("\n[COMPLETADO] El reporte 606 está listo.")
    if errors > 0:
        print(f"[AVISO] {errors} factura(s) tuvieron errores. Revisa la hoja 'Datos Extraídos' para más detalles.")


if __name__ == "__main__":
    main()
